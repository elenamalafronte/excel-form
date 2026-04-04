import time
import threading
from tkinter import BooleanVar, StringVar, filedialog, messagebox
from tkinter import font as tkfont
from pathlib import Path

from openpyxl import load_workbook

import config as cfg

from customtkinter import (
    CTkButton,
    CTkCheckBox,
    CTkComboBox,
    CTkEntry,
    CTkFont,
    CTkFrame,
    CTkLabel,
    CTkScrollableFrame,
    CTkTextbox,
    CTkToplevel,
)

from config import (
    COLUMNS,
    SEARCH_BY,
    get_next_fileNumber,
    get_next_fileNumber_from_value,
    save_columns_config,
    save_workbook_settings,
)
from excel import (
    _invalidate_desc_index_cache,
    append_row,
    get_description_for_itemcode,
    load_sheet,
    sync_form_sheet_columns,
)
from ui_style import (
    BODY_FONT_SIZE,
    BUTTON_CORNER_RADIUS,
    BUTTON_HEIGHT,
    CARD_CORNER_RADIUS,
    CONTROL_CORNER_RADIUS,
    ENTRY_HEIGHT,
    ENTRY_WIDTH,
    FORM_HEIGHT,
    FORM_WIDTH,
    LABEL_COLUMN_MIN_WIDTH,
    LABEL_FONT_SIZE,
    ROW_PADX,
    ROW_PADY,
    SECTION_TITLE_SIZE,
    TEXTBOX_HEIGHT,
)

# TODO: make font in table in search tab bigger (now too small)
# TODO: add dragger/something where you yourself can customise the width of the fields

def _open_file_picker(entry_widget):
    file_path = filedialog.askopenfilename(title="Select file for FileLink")
    if file_path:
        entry_widget.delete(0, "end")
        entry_widget.insert(0, file_path)


def _show_error(field_name, message):
    messagebox.showerror("Validation Error", f"{field_name}: {message}")


def _show_success(file_number):
    messagebox.showinfo("Saved", f"Row saved with File Number: {file_number}")


def _show_timed_success(file_number, elapsed_seconds, elapsed_before_save=None, elapsed_after_save=None):
    msg = f"Row saved with File Number: {file_number}"
    # optionally include save vs refresh breakdown for debugging performance issues
    # if elapsed_before_save is not None and elapsed_after_save is not None:
    #     save_duration = elapsed_after_save - elapsed_before_save
    #     refresh_duration = elapsed_seconds - elapsed_after_save
    #     msg += f"\n(Save: {save_duration:.3f}s, Refresh: {refresh_duration:.3f}s)"
    messagebox.showinfo("Saved", msg)


def _show_timed_error(message, elapsed_seconds):
    _show_error("Save", f"{message}\nElapsed: {elapsed_seconds:.3f}s")


_BUTTON_IDLE_FG = ("#3B8ED0", "#1F6AA5")
_BUTTON_IDLE_HOVER = ("#36719F", "#144870")
_BUTTON_BUSY_FG = "#1F4D82"


def _set_button_saving_state(button, is_saving, idle_text, busy_text="Saving...", refresh_widget=None):
    if button is None:
        return

    if is_saving:
        button.configure(
            state="normal",
            text=busy_text,
            fg_color=_BUTTON_BUSY_FG,
            hover_color=_BUTTON_BUSY_FG,
        )
        if refresh_widget is not None:
            refresh_widget.update_idletasks()
    else:
        button.configure(
            state="normal",
            text=idle_text,
            fg_color=_BUTTON_IDLE_FG,
            hover_color=_BUTTON_IDLE_HOVER,
        )


_sheet_rows_cache = None
_sheet_rows_cache_file_sig = None


def _invalidate_sheet_rows_cache():
    global _sheet_rows_cache, _sheet_rows_cache_file_sig
    _sheet_rows_cache = None
    _sheet_rows_cache_file_sig = None


def _get_cached_sheet_rows():
    """Return workbook rows for autofill, reloading only when the file changes.

    TODO: if you later want even faster autofill, keep this cache warm after
    startup instead of waiting for the first ItemCode lookup.
    """
    global _sheet_rows_cache, _sheet_rows_cache_file_sig

    workbook_path = Path(cfg.EXCEL_FILE)
    if not workbook_path.exists():
        _sheet_rows_cache = []
        _sheet_rows_cache_file_sig = None
        return _sheet_rows_cache

    try:
        stat = workbook_path.stat()
        current_file_sig = (int(stat.st_mtime_ns), int(stat.st_size))
    except OSError:
        return load_sheet()

    if current_file_sig != _sheet_rows_cache_file_sig:
        _sheet_rows_cache = load_sheet()
        _sheet_rows_cache_file_sig = current_file_sig

    return _sheet_rows_cache


def _lookup_description_for_itemcode(item_code, rows):
    """Return the description for an ItemCode.

    This is the one place where the autofill rule lives.

    If your workbook ever changes shape, update the matching logic here and
    leave the rest of the UI code alone.
    """
    item_code = str(item_code).strip().upper()

    for row in rows or []:
        if str(row.get("ItemCode", "")).strip().upper() == item_code:
            # Prefer a non-empty description when duplicates exist.
            description = (row.get("Description", "") or "").strip()
            if description:
                return description

    # Fallback to source sheet lookup so new ItemCodes autofill before save.
    return get_description_for_itemcode(item_code)


def _update_description_field(description_widget, item_code_widget):
    """Update the Description field from the current ItemCode value.

    This is intentionally small: get the current ItemCode, look up the
    description, and write it into the disabled field.
    """
    item_code = item_code_widget.get().strip()

    if not item_code:
        description_widget.configure(state="normal")
        description_widget.delete("1.0", "end")
        description_widget.configure(state="disabled")
        _autosize_description_widget(description_widget)
        return

    rows = _get_cached_sheet_rows()
    description = _lookup_description_for_itemcode(item_code, rows)

    description_widget.configure(state="normal")
    description_widget.delete("1.0", "end")
    description_widget.insert("1.0", description)
    description_widget.configure(state="disabled")
    _autosize_description_widget(description_widget)


def _autosize_description_widget(description_widget, min_height=TEXTBOX_HEIGHT + 10):
    """Grow the description textbox only when wrapped text would overflow.

    Keeps a stable baseline height and expands to fit the number of displayed
    wrapped lines currently needed by the text widget.
    """
    try:
        description_widget.update_idletasks()
        display_lines = description_widget.count("1.0", "end-1c", "displaylines")
        line_count = int(display_lines[0]) if display_lines else 1
    except Exception:
        line_count = 1

    try:
        body_font = tkfont.Font(font=description_widget.cget("font"))
        line_height = max(1, int(body_font.metrics("linespace")))
    except Exception:
        line_height = 18

    # Include internal text widget padding/border so the last line is visible.
    target_height = max(min_height, (line_count * line_height) + 14)
    description_widget.configure(height=target_height)


def _bind_itemcode_autofill(item_code_widget, description_widget):
    """Wire ItemCode edits to description autofill.
    """
    def _on_itemcode_event(event=None):
        _update_description_field(description_widget, item_code_widget)

    item_code_widget.bind("<FocusOut>", _on_itemcode_event)

    # make autofill to run on every keystroke.
    item_code_widget.bind("<KeyRelease>", _on_itemcode_event)

    return _on_itemcode_event


def _validate_workbook_settings(excel_file, source_sheet_name, form_sheet_name):
    workbook_path = Path(excel_file).expanduser()
    if not workbook_path.exists():
        return False, f"Workbook not found: {workbook_path}"

    try:
        wb = load_workbook(workbook_path, read_only=True)
        sheetnames = list(wb.sheetnames)
        wb.close()
    except Exception as exc:
        return False, f"Could not open workbook:\n{exc}"

    missing = []
    if source_sheet_name not in sheetnames:
        missing.append(f"source sheet '{source_sheet_name}'")
    if form_sheet_name not in sheetnames:
        missing.append(f"form sheet '{form_sheet_name}'")

    if missing:
        return False, "Workbook is missing: " + ", ".join(missing)

    return True, ""


def build_insert_tab(tab):
    label_font = CTkFont(size=LABEL_FONT_SIZE)
    body_font = CTkFont(size=BODY_FONT_SIZE)
    title_font = CTkFont(size=SECTION_TITLE_SIZE, weight="bold")
    field_types = ["text", "number", "general", "filelink"]
    existing_by_name = {c.get("name"): c for c in COLUMNS}

    def _rebuild_insert_tab():
        for child in tab.winfo_children():
            child.destroy()
        build_insert_tab(tab)

    def _open_workbook_settings():
        panel = CTkToplevel(tab)
        panel.title("Workbook Settings")
        panel.geometry("680x260")
        panel.transient(tab.winfo_toplevel())
        panel.grab_set()

        CTkLabel(panel, text="Workbook Settings", font=CTkFont(size=18, weight="bold")).pack(
            anchor="w", padx=12, pady=(12, 8)
        )

        form = CTkFrame(panel, fg_color="transparent")
        form.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        form.grid_columnconfigure(1, weight=1)

        workbook_var = StringVar(value=cfg.EXCEL_FILE)
        source_sheet_var = StringVar(value=cfg.SOURCE_SHEET_NAME)
        form_sheet_var = StringVar(value=cfg.FORM_SHEET_NAME)

        def _browse_workbook():
            selected = filedialog.askopenfilename(
                title="Select Workbook",
                filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
            )
            if selected:
                workbook_var.set(selected)

        def _save_workbook_settings():
            workbook_path = workbook_var.get().strip()
            source_sheet_name = source_sheet_var.get().strip()
            form_sheet_name = form_sheet_var.get().strip()

            if not workbook_path:
                messagebox.showerror("Workbook Settings", "Workbook file is required.")
                return
            if not source_sheet_name:
                messagebox.showerror("Workbook Settings", "Source sheet name is required.")
                return
            if not form_sheet_name:
                messagebox.showerror("Workbook Settings", "Form sheet name is required.")
                return

            is_valid, message = _validate_workbook_settings(workbook_path, source_sheet_name, form_sheet_name)
            if not is_valid:
                messagebox.showerror("Workbook Settings", message)
                return

            try:
                save_workbook_settings(
                    excel_file=str(Path(workbook_path).expanduser().resolve()),
                    source_sheet_name=source_sheet_name,
                    form_sheet_name=form_sheet_name,
                )
            except Exception as exc:
                messagebox.showerror("Workbook Settings", f"Could not save workbook settings:\n{exc}")
                return

            _invalidate_sheet_rows_cache()
            _invalidate_desc_index_cache()

            panel.destroy()
            _rebuild_insert_tab()
            messagebox.showinfo("Workbook Settings", "Workbook settings saved and applied.")

        CTkLabel(form, text="Workbook File", font=label_font).grid(row=0, column=0, sticky="w", padx=(0, 10), pady=8)
        workbook_entry = CTkEntry(form, font=body_font, textvariable=workbook_var)
        workbook_entry.grid(row=0, column=1, sticky="ew", padx=(0, 10), pady=8)
        CTkButton(form, text="Browse", width=90, command=_browse_workbook).grid(row=0, column=2, sticky="e", pady=8)

        CTkLabel(form, text="Source Sheet", font=label_font).grid(row=1, column=0, sticky="w", padx=(0, 10), pady=8)
        source_entry = CTkEntry(form, font=body_font, textvariable=source_sheet_var)
        source_entry.grid(row=1, column=1, columnspan=2, sticky="ew", pady=8)

        CTkLabel(form, text="Form Sheet", font=label_font).grid(row=2, column=0, sticky="w", padx=(0, 10), pady=8)
        form_entry = CTkEntry(form, font=body_font, textvariable=form_sheet_var)
        form_entry.grid(row=2, column=1, columnspan=2, sticky="ew", pady=8)

        footer = CTkFrame(panel, fg_color="transparent")
        footer.pack(fill="x", padx=12, pady=(0, 12))
        CTkButton(footer, text="Cancel", width=100, command=panel.destroy).pack(side="right")
        CTkButton(
            footer,
            text="Save",
            width=100,
            command=_save_workbook_settings,
        ).pack(side="right", padx=(0, 8))

    def _open_fields_customizer():
        panel = CTkToplevel(tab)
        panel.title("Customize Fields")
        panel.geometry("720x560")
        panel.transient(tab.winfo_toplevel())
        panel.grab_set()

        CTkLabel(panel, text="Customize Insert Fields", font=CTkFont(size=18, weight="bold")).pack(
            anchor="w", padx=12, pady=(12, 8)
        )

        rows_frame = CTkScrollableFrame(panel)
        rows_frame.pack(fill="both", expand=True, padx=12, pady=(0, 10))
        rows_frame.grid_columnconfigure(0, weight=0)
        rows_frame.grid_columnconfigure(1, weight=3)
        rows_frame.grid_columnconfigure(2, weight=2)
        rows_frame.grid_columnconfigure(3, weight=0)
        required_lane_width = 76
        checkbox_size = 22
        checkbox_pad_x = max((required_lane_width - checkbox_size) // 2, 0)

        CTkLabel(rows_frame, text="Field Name", font=CTkFont(size=13, weight="bold")).grid(
            row=0, column=1, sticky="w", padx=8, pady=(8, 6)
        )
        CTkLabel(rows_frame, text="Type", font=CTkFont(size=13, weight="bold")).grid(
            row=0, column=2, sticky="w", padx=8, pady=(8, 6)
        )
        CTkLabel(
            rows_frame,
            text="Required",
            font=CTkFont(size=13, weight="bold"),
            width=required_lane_width,
            anchor="center",
        ).grid(row=0, column=3, sticky="w", padx=(0, 0), pady=(8, 6))

        row_models = []
        removed_rows_stack = []
        undo_button = None
        save_button = None
        save_in_progress = {"value": False}
        drag_state = {"model": None}

        def _set_save_button_saving_state(is_saving):
            if save_button is None:
                return
            if is_saving:
                save_in_progress["value"] = True
                _set_button_saving_state(
                    save_button,
                    True,
                    idle_text="Save",
                    busy_text="Saving...",
                    refresh_widget=panel,
                )
            else:
                save_in_progress["value"] = False
                _set_button_saving_state(save_button, False, idle_text="Save")

        def _update_undo_button_state():
            if undo_button is None:
                return
            undo_button.configure(state="normal" if removed_rows_stack else "disabled")

        def _reflow_rows():
            for idx, row in enumerate(row_models, start=1):
                row["drag_handle"].grid(row=idx, column=0, sticky="w", padx=(8, 4), pady=4)
                row["name_entry"].grid(row=idx, column=1, sticky="ew", padx=8, pady=4)
                row["type_combo"].grid(row=idx, column=2, sticky="ew", padx=8, pady=4)
                row["action_cell"].grid(row=idx, column=3, sticky="w", padx=(6, 6), pady=4)

        def _move_row(model, target_index):
            if model not in row_models:
                return
            current_index = row_models.index(model)
            target_index = max(0, min(target_index, len(row_models) - 1))
            if current_index == target_index:
                return
            row_models.pop(current_index)
            row_models.insert(target_index, model)
            _reflow_rows()

        def _drag_target_index():
            pointer_y = rows_frame.winfo_pointery()
            for idx, row in enumerate(row_models):
                center_y = row["name_entry"].winfo_rooty() + (row["name_entry"].winfo_height() // 2)
                if pointer_y < center_y:
                    return idx
            return len(row_models) - 1

        def _on_drag_start(event, model):
            drag_state["model"] = model
            handle = model.get("drag_handle")
            if handle is not None:
                handle.configure(fg_color="#2C8C67")

        def _on_drag_motion(event):
            model = drag_state.get("model")
            if model is None:
                return
            _move_row(model, _drag_target_index())

        def _on_drag_end(event):
            model = drag_state.get("model")
            if model is not None:
                handle = model.get("drag_handle")
                if handle is not None:
                    handle.configure(fg_color="transparent")
            drag_state["model"] = None

        def _remove_row(model):
            if len(row_models) <= 1:
                messagebox.showwarning("Customize Fields", "At least one field is required.")
                return

            row_index = row_models.index(model)
            removed_field = {
                "name": model["name_entry"].get().strip(),
                "type": model["type_combo"].get().strip().lower() or "text",
                "required": bool(model["required_var"].get()),
                "original_name": model.get("original_name"),
            }

            for widget in (
                model["name_entry"],
                model["type_combo"],
                model["drag_handle"],
                model["action_cell"],
            ):
                widget.destroy()

            row_models.remove(model)
            removed_rows_stack.append({"index": row_index, "field": removed_field})
            _reflow_rows()
            _update_undo_button_state()

        def _undo_remove_row():
            if not removed_rows_stack:
                return

            removed_entry = removed_rows_stack.pop()
            _add_row(
                initial=removed_entry["field"],
                insert_at=min(max(0, removed_entry["index"]), len(row_models)),
            )
            _update_undo_button_state()

        def _add_row(initial=None, insert_at=None):
            initial = initial or {}
            required_var = BooleanVar(value=bool(initial.get("required", False)))

            name_entry = CTkEntry(rows_frame)
            name_entry.insert(0, str(initial.get("name", "")))

            type_combo = CTkComboBox(rows_frame, values=field_types)
            initial_type = str(initial.get("type", "text")).strip().lower()
            type_combo.set(initial_type if initial_type in field_types else "text")

            action_cell = CTkFrame(rows_frame, fg_color="transparent")

            required_check = CTkCheckBox(
                action_cell,
                text="",
                variable=required_var,
                width=26,
                checkbox_width=checkbox_size,
                checkbox_height=checkbox_size,
            )
            required_check.pack(side="left", padx=(checkbox_pad_x, checkbox_pad_x))

            drag_handle = CTkLabel(
                rows_frame,
                text="::",
                width=26,
                corner_radius=6,
                fg_color="transparent",
                cursor="fleur",
            )

            model = {
                "original_name": initial.get("original_name", initial.get("name")),
                "name_entry": name_entry,
                "type_combo": type_combo,
                "required_var": required_var,
                "required_check": required_check,
                "action_cell": action_cell,
                "drag_handle": drag_handle,
                "remove_btn": None,
            }

            drag_handle.bind("<ButtonPress-1>", lambda event, m=model: _on_drag_start(event, m))
            drag_handle.bind("<B1-Motion>", _on_drag_motion)
            drag_handle.bind("<ButtonRelease-1>", _on_drag_end)

            remove_btn = CTkButton(
                action_cell,
                text="Remove",
                width=86,
                command=lambda m=model: _remove_row(m),
            )
            remove_btn.pack(side="left", padx=(12, 0))
            model["remove_btn"] = remove_btn

            if insert_at is None or insert_at < 0 or insert_at > len(row_models):
                row_models.append(model)
            else:
                row_models.insert(insert_at, model)
            _reflow_rows()

        def _save_customized_fields():
            if save_in_progress["value"]:
                return
            _set_save_button_saving_state(True)

            old_columns_snapshot = [dict(c) for c in COLUMNS]
            new_columns = []
            seen_names = set()

            for row in row_models:
                name = row["name_entry"].get().strip()
                col_type = row["type_combo"].get().strip().lower()
                required = bool(row["required_var"].get())

                if not name:
                    _set_save_button_saving_state(False)
                    messagebox.showerror("Customize Fields", "Field name cannot be empty.")
                    return
                if name in seen_names:
                    _set_save_button_saving_state(False)
                    messagebox.showerror("Customize Fields", f"Duplicate field name: {name}")
                    return
                if col_type not in field_types:
                    _set_save_button_saving_state(False)
                    messagebox.showerror("Customize Fields", f"Invalid type for {name}: {col_type}")
                    return

                seen_names.add(name)

                original_name = row.get("original_name")
                existing = existing_by_name.get(original_name) or existing_by_name.get(name) or {}
                col_def = {
                    "name": name,
                    "type": col_type,
                    "required": required,
                }

                if "unique" in existing:
                    col_def["unique"] = existing["unique"]
                if "validate" in existing:
                    col_def["validate"] = existing["validate"]

                if name == "File Number":
                    col_def["required"] = True
                    col_def["unique"] = True
                    col_def["validate"] = "is_valid_fileNumber"

                new_columns.append(col_def)

            if not new_columns:
                _set_save_button_saving_state(False)
                messagebox.showerror("Customize Fields", "At least one field is required.")
                return

            has_item_code = any(col.get("name") == "ItemCode" for col in new_columns)
            has_description = any(col.get("name") == "Description" for col in new_columns)
            if has_description and not has_item_code:
                _set_save_button_saving_state(False)
                messagebox.showerror(
                    "Customize Fields",
                    "Description requires ItemCode so the auto-formula can keep working.",
                )
                return

            try:
                save_columns_config(new_columns)
            except Exception as exc:
                _set_save_button_saving_state(False)
                messagebox.showerror("Customize Fields", f"Could not save config.py:\n{exc}")
                return

            try:
                sync_form_sheet_columns(old_columns_snapshot, new_columns)
            except Exception as exc:
                # Roll back config file so app config and workbook schema do not diverge.
                try:
                    save_columns_config(old_columns_snapshot)
                except Exception:
                    pass
                _set_save_button_saving_state(False)
                messagebox.showerror(
                    "Customize Fields",
                    "Config was not applied because workbook schema sync failed.\n"
                    f"Reason: {exc}",
                )
                return

            COLUMNS.clear()
            COLUMNS.extend(new_columns)

            _invalidate_sheet_rows_cache()

            SEARCH_BY.clear()
            SEARCH_BY.extend([col["name"] for col in new_columns])

            rebuild_search = getattr(tab, "rebuild_search", None)
            if callable(rebuild_search):
                rebuild_search()

            panel.destroy()
            messagebox.showinfo(
                "Customize Fields",
                "Field configuration saved and workbook columns updated.",
            )
            _rebuild_insert_tab()

        for col in COLUMNS:
            _add_row(col)

        buttons = CTkFrame(panel, fg_color="transparent")
        buttons.pack(fill="x", padx=12, pady=(0, 12))

        CTkButton(buttons, text="Add Field", width=100, command=lambda: _add_row()).pack(side="left")
        undo_button = CTkButton(buttons, text="Undo Remove", width=120, command=_undo_remove_row)
        undo_button.pack(side="left", padx=(8, 0))
        _update_undo_button_state()
        CTkButton(buttons, text="Cancel", width=100, command=panel.destroy).pack(side="right")
        save_button = CTkButton(buttons, text="Save", width=100, command=_save_customized_fields)
        save_button.pack(side="right", padx=(0, 8))

    outer_container = CTkFrame(tab, fg_color="transparent")
    outer_container.pack(fill="both", expand=True, padx=12, pady=12)

    # Centered, scrollable card to keep the insert form tidy and readable.
    container = CTkScrollableFrame(
        outer_container,
        width=FORM_WIDTH,
        height=FORM_HEIGHT,
        corner_radius=CARD_CORNER_RADIUS,
    )
    container.pack(pady=8)
    container.grid_columnconfigure(0, minsize=LABEL_COLUMN_MIN_WIDTH)

    header_row = CTkFrame(container, fg_color="transparent")
    header_row.grid(row=0, column=0, columnspan=3, sticky="ew", padx=ROW_PADX, pady=(10, 6))
    header_row.grid_columnconfigure(0, weight=1)
    header_row.grid_columnconfigure(1, weight=0)
    header_row.grid_columnconfigure(2, weight=0)

    CTkLabel(header_row, text="Insert Material Record", font=title_font).grid(
        row=0,
        column=0,
        sticky="w",
    )
    CTkButton(
        header_row,
        text="Customize Fields",
        height=BUTTON_HEIGHT,
        corner_radius=BUTTON_CORNER_RADIUS,
        font=body_font,
        command=_open_fields_customizer,
    ).grid(row=0, column=1, sticky="e")
    CTkButton(
        header_row,
        text="Workbook Settings",
        height=BUTTON_HEIGHT,
        corner_radius=BUTTON_CORNER_RADIUS,
        font=body_font,
        command=_open_workbook_settings,
    ).grid(row=0, column=2, sticky="e", padx=(8, 0))

    next_file_number_state = {"value": None}

    try:
        next_file_number_state["value"] = get_next_fileNumber(load_sheet())
    except Exception:
        next_file_number_state["value"] = "01-01A"

    fields = {}
    item_code_widget = None
    description_widget = None
    for row_idx, col in enumerate(COLUMNS, start=1):
        filelink_header = None
        if col.get("type") == "filelink":
            filelink_header = CTkFrame(container, fg_color="transparent")
            filelink_header.grid(
                row=row_idx,
                column=0,
                sticky="ew",
                padx=ROW_PADX,
                pady=ROW_PADY,
            )
            filelink_header.grid_columnconfigure(0, weight=1)
            CTkLabel(filelink_header, text=col["name"], font=label_font).grid(
                row=0,
                column=0,
                sticky="w",
            )
        else:
            CTkLabel(container, text=col["name"], font=label_font).grid(
                row=row_idx,
                column=0,
                sticky="w",
                padx=ROW_PADX,
                pady=ROW_PADY,
            )

        if col["name"] == "File Number":
            widget = CTkEntry(
                container,
                width=ENTRY_WIDTH,
                height=ENTRY_HEIGHT,
                corner_radius=CONTROL_CORNER_RADIUS,
                font=body_font,
            )
            widget.insert(0, "Auto-generated on Save")
            widget.configure(state="disabled")
            widget.grid(row=row_idx, column=1, sticky="ew", padx=ROW_PADX, pady=ROW_PADY)
        elif col["name"] == "Description":
            widget = CTkTextbox(
                container,
                width=ENTRY_WIDTH,
                height=TEXTBOX_HEIGHT + 10,  # add a bit of extra height so its visually clear this is a multi-line field
                corner_radius=CONTROL_CORNER_RADIUS,
                font=body_font,
                border_width=2,
                border_color="#979DA2",
            )
            widget.configure(state="disabled")
            widget.grid(row=row_idx, column=1, sticky="ew", padx=ROW_PADX, pady=ROW_PADY)
            widget.bind("<Configure>", lambda event, w=widget: _autosize_description_widget(w), add="+")
            description_widget = widget
        elif col["name"] == "ItemCode":
            widget = CTkEntry(
                container,
                width=ENTRY_WIDTH,
                height=ENTRY_HEIGHT,
                corner_radius=CONTROL_CORNER_RADIUS,
                font=body_font,
            )
            widget.grid(row=row_idx, column=1, sticky="ew", padx=ROW_PADX, pady=ROW_PADY)
            fields["ItemCode"] = widget
            item_code_widget = widget
        elif col.get("type") == "filelink":
            widget = CTkEntry(
                container,
                width=ENTRY_WIDTH,
                height=ENTRY_HEIGHT,
                corner_radius=CONTROL_CORNER_RADIUS,
                font=body_font,
            )
            CTkButton(
                filelink_header,
                text="Browse",
                width=100,
                height=BUTTON_HEIGHT,
                corner_radius=BUTTON_CORNER_RADIUS,
                font=body_font,
                command=lambda w=widget: _open_file_picker(w),
            ).grid(row=0, column=1, sticky="e")
            widget.grid(row=row_idx, column=1, sticky="ew", padx=ROW_PADX, pady=ROW_PADY)
        else:
            widget = CTkEntry(
                container,
                width=ENTRY_WIDTH,
                height=ENTRY_HEIGHT,
                corner_radius=CONTROL_CORNER_RADIUS,
                font=body_font,
            )
            widget.grid(row=row_idx, column=1, sticky="ew", padx=ROW_PADX, pady=ROW_PADY)

        fields[col["name"]] = widget

    if item_code_widget is not None and description_widget is not None:
        _bind_itemcode_autofill(item_code_widget, description_widget)
        _autosize_description_widget(description_widget)

    # keep a reference so the thread callback can re-enable it
    save_button = CTkButton(
        container,
        text="Save Row",
        height=BUTTON_HEIGHT,
        corner_radius=BUTTON_CORNER_RADIUS,
        font=body_font,
    )
    save_row_in_progress = {"value": False}

    def on_submit():
        if save_row_in_progress["value"]:
            return

        started_at = time.perf_counter()
        file_number = next_file_number_state["value"]
        if not file_number:
            _show_error("File Number", "Could not determine the next File Number")
            return

        data = {}
        for col in COLUMNS:
            name = col["name"]
            if name == "File Number":
                data[name] = file_number
                continue
            if name == "Description":
                # The insert tab owns this value now.
                # If the ItemCode lookup has not run yet, this may still be blank.
                data[name] = fields[name].get("1.0", "end").strip()
                continue

            value = fields[name].get().strip()
            if col.get("required") and not value:
                _show_error(name, "This field is required")
                return
            data[name] = value

        # keep button dark while saving; prevent double-submit with explicit guard.
        save_row_in_progress["value"] = True
        _set_button_saving_state(
            save_button,
            True,
            idle_text="Save Row",
            busy_text="Saving...",
            refresh_widget=tab,
        )

        def do_save():
            try:
                elapsed_before_save = time.perf_counter() - started_at
                append_row(data)
                elapsed_after_save = time.perf_counter() - started_at
                tab.after(0, _on_save_success, file_number, started_at, elapsed_before_save, elapsed_after_save)
            except Exception as exc:
                tab.after(0, _on_save_error, str(exc), started_at)

        threading.Thread(target=do_save, daemon=True).start()

    def _on_save_success(file_number, started_at, elapsed_before_save, elapsed_after_save):
        elapsed_seconds = time.perf_counter() - started_at
        save_row_in_progress["value"] = False
        _set_button_saving_state(save_button, False, idle_text="Save Row")

        # Ensure immediate ItemCode autofill uses fresh workbook rows.
        _invalidate_sheet_rows_cache()

        _show_timed_success(file_number, elapsed_seconds, elapsed_before_save, elapsed_after_save)
        try:
            next_file_number_state["value"] = get_next_fileNumber_from_value(file_number)
        except Exception:
            next_file_number_state["value"] = None

        for col in COLUMNS:
            # Skip auto-filled fields. They are managed separately and should not
            # be cleared here unless you explicitly want to reset them after save.
            if col["name"] == "File Number":
                continue
            if col["name"] == "Description":
                fields[col["name"]].configure(state="normal")
                fields[col["name"]].delete("1.0", "end")
                fields[col["name"]].configure(state="disabled")
                _autosize_description_widget(fields[col["name"]])
            else:
                fields[col["name"]].delete(0, "end")

        refresh_search = getattr(tab, "refresh_search", None)
        if callable(refresh_search):
            refresh_search()

    def _on_save_error(message, started_at):
        elapsed_seconds = time.perf_counter() - started_at
        save_row_in_progress["value"] = False
        _set_button_saving_state(save_button, False, idle_text="Save Row")
        _show_timed_error(message, elapsed_seconds)

    save_button.configure(command=on_submit)
    save_button.grid(
        row=len(COLUMNS) + 2,
        column=0,
        columnspan=3,
        sticky="ew",
        padx=ROW_PADX,
        pady=(12, 16),
    )
