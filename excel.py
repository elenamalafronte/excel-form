import os
import subprocess
import zipfile
import re
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

import config as cfg

from config import (
    COLUMNS,
    build_description_formula,
    FILE_NUMBER_PATTERN,
    SEARCH_BY,
)


APP_DATA_START_ROW = 3
EMPTY_ROW_STOP_THRESHOLD = 25


def _open_workbook():
    path = Path(cfg.EXCEL_FILE)
    if path.exists():
        return load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm")
    return Workbook()


def _get_layout_sheets(wb):
    ws_source = None
    ws_form = None

    if cfg.SOURCE_SHEET_NAME in wb.sheetnames:
        ws_source = wb[cfg.SOURCE_SHEET_NAME]
    elif len(wb.worksheets) > 0:
        ws_source = wb.worksheets[0]

    if cfg.FORM_SHEET_NAME in wb.sheetnames:
        ws_form = wb[cfg.FORM_SHEET_NAME]
    elif len(wb.worksheets) > 1:
        ws_form = wb.worksheets[1]

    if ws_source is None:
        ws_source = wb.active
        ws_source.title = cfg.SOURCE_SHEET_NAME
    if ws_form is None:
        ws_form = wb.create_sheet(title=cfg.FORM_SHEET_NAME)

    return ws_source, ws_form


def _get_form_sheet_for_read(wb):
    if cfg.FORM_SHEET_NAME in wb.sheetnames:
        return wb[cfg.FORM_SHEET_NAME]
    if len(wb.worksheets) < 2:
        return None
    return wb.worksheets[1]


def _ensure_workbook_and_sheets():
    wb = _open_workbook()
    ws_source, ws_form = _get_layout_sheets(wb)

    headers = [c["name"] for c in COLUMNS]
    if ws_form.max_row == 1 and ws_form.cell(row=1, column=1).value is None:
        ws_form.append(headers)
    else:
        for idx, header in enumerate(headers, start=1):
            if ws_form.cell(row=1, column=idx).value != header:
                ws_form.cell(row=1, column=idx, value=header)

    return wb, ws_source, ws_form


def _col_idx_to_letter(col_idx):
    """Convert 0-based column index to Excel column letter (A, B, ..., AA, ...)."""
    letters = []
    n = col_idx + 1
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters.append(chr(65 + rem))
    return "".join(reversed(letters))


def _find_sheet_zip_path(file_path, sheet_name):
    """Return the ZIP-internal path to a named worksheet's XML file."""
    with zipfile.ZipFile(file_path, "r") as z:
        wb_xml = z.read("xl/workbook.xml").decode("utf-8")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    id_match = re.search(
        rf'<sheet\b[^>]*\bname="{re.escape(sheet_name)}"[^>]*\br:id="([^"]+)"', wb_xml
    ) or re.search(
        rf'<sheet\b[^>]*\br:id="([^"]+)"[^>]*\bname="{re.escape(sheet_name)}"', wb_xml
    )
    if not id_match:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.xml")
    r_id = id_match.group(1)

    target_match = re.search(
        rf'<Relationship\b[^>]*\bId="{re.escape(r_id)}"[^>]*\bTarget="([^"]+)"', rels_xml
    ) or re.search(
        rf'<Relationship\b[^>]*\bTarget="([^"]+)"[^>]*\bId="{re.escape(r_id)}"', rels_xml
    )
    if not target_match:
        raise ValueError(f"Could not resolve relationship {r_id}")
    target = target_match.group(1)

    # Normalize relationship target to actual ZIP member format.
    # Some workbooks store absolute-like targets such as /xl/worksheets/sheet2.xml
    # while ZIP members are stored without leading slash.
    normalized = target.replace("\\", "/").lstrip("/")
    if not normalized.startswith("xl/"):
        normalized = f"xl/{normalized}"
    return normalized


def _get_sheet_row_bounds(file_path: Path, sheet_name: str) -> tuple[int, int]:
    """Return (first_row_with_tag, last_row_with_tag) for a worksheet XML.

    Works across different workbook layouts where visual headers/data may start
    far below row 1 because of formatting blocks.
    """
    sheet_zip_path = _find_sheet_zip_path(file_path, sheet_name)
    with zipfile.ZipFile(file_path, "r") as z:
        sheet_str = z.read(sheet_zip_path).decode("utf-8", "ignore")

    row_numbers = [int(x) for x in re.findall(r'<row[^>]*\sr="(\d+)"', sheet_str)]
    if not row_numbers:
        return 1, 1
    return min(row_numbers), max(row_numbers)


def _build_row_xml(row_idx, row_values, cached_values=None):
    """Return the XML string for one worksheet row, using inline strings.

    cached_values: optional {col_idx: value} — for formula cells, the pre-computed
    result is written as <v> so openpyxl data_only=True can read it immediately
    without waiting for Excel to recalculate.
    """
    cells = []
    cached_values = cached_values or {}
    for col_idx, value in enumerate(row_values):
        if value is None or value == "":
            continue
        col_letter = _col_idx_to_letter(col_idx)
        cell_ref = f"{col_letter}{row_idx}"
        val = str(value)
        if val.startswith("="):
            formula = val[1:].replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            cached = cached_values.get(col_idx)
            if cached is not None and cached != "":
                safe_v = str(cached).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                cells.append(f'<c r="{cell_ref}" t="str"><f>{formula}</f><v>{safe_v}</v></c>')
            else:
                cells.append(f'<c r="{cell_ref}"><f>{formula}</f></c>')
        else:
            try:
                num = float(val)
                stored = int(num) if num == int(num) and "." not in val else num
                cells.append(f'<c r="{cell_ref}"><v>{stored}</v></c>')
            except ValueError:
                safe = val.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                cells.append(f'<c r="{cell_ref}" t="inlineStr"><is><t>{safe}</t></is></c>')
    return f'<row r="{row_idx}">{"".join(cells)}</row>'


def _zip_append_row(file_path, row_values, new_row_idx, cached_values=None):
    """Append one row by rewriting only the Heat Number worksheet XML in the ZIP.

    CREXPD01's bytes are copied verbatim — it is never parsed or serialized.
    This eliminates the save latency that came from openpyxl re-serializing
    15 000+ rows of source data on every insert.
    """
    sheet_zip_path = _find_sheet_zip_path(file_path, cfg.FORM_SHEET_NAME)
    new_row_xml = _build_row_xml(new_row_idx, row_values, cached_values)

    with zipfile.ZipFile(file_path, "r") as z:
        sheet_str = z.read(sheet_zip_path).decode("utf-8")

    existing_row_full = re.compile(rf'<row[^>]*\br="{new_row_idx}"[^>]*>.*?</row>', re.S)
    existing_row_self = re.compile(rf'<row[^>]*\br="{new_row_idx}"[^>]*/>', re.S)

    if existing_row_full.search(sheet_str):
        sheet_str = existing_row_full.sub(new_row_xml, sheet_str, count=1)
    elif existing_row_self.search(sheet_str):
        sheet_str = existing_row_self.sub(new_row_xml, sheet_str, count=1)
    elif "</sheetData>" in sheet_str:
        sheet_str = sheet_str.replace("</sheetData>", f"{new_row_xml}</sheetData>", 1)
    elif re.search(r"<sheetData\s*/>", sheet_str):
        # Empty sheets can use a self-closing sheetData tag.
        sheet_str = re.sub(
            r"<sheetData\s*/>",
            f"<sheetData>{new_row_xml}</sheetData>",
            sheet_str,
            count=1,
        )
    else:
        raise ValueError(f"sheetData section not found in {sheet_zip_path}")
    def _update_dimension(match):
        current_max = int(match.group(3))
        max_row = max(current_max, new_row_idx)
        return f"{match.group(1)}{match.group(2)}{max_row}{match.group(4)}"

    sheet_str = re.sub(
        r'(<dimension ref="[A-Z]+\d+:)([A-Z]+)(\d+)(")',
        _update_dimension,
        sheet_str,
    )

    tmp_path = file_path.with_suffix(".xlsx.tmp")
    try:
        with zipfile.ZipFile(file_path, "r") as z_in, \
             zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as z_out:
            for item in z_in.infolist():
                data = sheet_str.encode("utf-8") if item.filename == sheet_zip_path \
                    else z_in.read(item.filename)
                z_out.writestr(item, data)
        
        # Handle Windows file locking by retrying replace if it fails
        max_retries = 3
        for attempt in range(max_retries):
            try:
                tmp_path.replace(file_path)
                break
            except (OSError, PermissionError) as e:
                if attempt < max_retries - 1:
                    time.sleep(0.1)  # Brief delay before retry
                else:
                    raise
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise


# Column O (1-based = 15, 0-based = 14) is the ItemCode column in CREXPD01,
# as confirmed by the formula: MATCH(B{row}, CREXPD01!$O$1:$O$15000, 0)
_CREXPD01_ITEMCODE_COL_FALLBACK = 14  # 0-based index


def _col_str_to_idx(col_str: str) -> int:
    """Convert a column letter string like 'A' or 'AV' to a 0-based index."""
    n = 0
    for ch in col_str:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def _build_description_index(file_path: Path) -> dict:
    """Return {ItemCode.upper(): description} from source sheet.

    Reads from Ident Code (column O) and Detailed Description (column R) columns.
    Uses openpyxl with direct column detection to match the Excel formula mapping.
    """
    return _build_description_index_fallback(file_path)



def _build_description_index_fallback(file_path: Path) -> dict:
    """Fallback description index builder using openpyxl row values."""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        if cfg.SOURCE_SHEET_NAME in wb.sheetnames:
            ws = wb[cfg.SOURCE_SHEET_NAME]
        elif wb.worksheets:
            ws = wb.worksheets[0]
        else:
            wb.close()
            return {}

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            wb.close()
            return {}

        normalized_headers = [
            str(h or "").strip().lower().replace(" ", "").replace("_", "")
            for h in header_row
        ]

        item_idx = -1
        desc_idx = -1
        for idx, name in enumerate(normalized_headers):
            # Prefer "Ident Code" over "Item Code" to match Excel formula that uses column O
            if "identcode" in name:
                item_idx = idx
            elif "itemcode" in name and item_idx < 0:
                item_idx = idx
            if desc_idx < 0 and "detaileddescription" in name:
                desc_idx = idx

        if item_idx < 0 or desc_idx < 0:
            wb.close()
            return {}

        index = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            item_code = ""
            description = ""

            if item_idx < len(row) and row[item_idx] is not None:
                item_code = str(row[item_idx]).strip().upper()
            if desc_idx < len(row) and row[desc_idx] is not None:
                description = str(row[desc_idx]).strip()

            if item_code:
                index[item_code] = description

        wb.close()
        return index
    except Exception:
        return {}


# Module-level description index cache.
# Rebuilt only when workbook settings change (or cache is explicitly invalidated),
# so saving a form row does not trigger a full source-sheet rescan.
_desc_index_cache: dict = {}
_desc_index_cache_key: tuple[str, str] | None = None


def _invalidate_desc_index_cache() -> None:
    global _desc_index_cache, _desc_index_cache_key
    _desc_index_cache = {}
    _desc_index_cache_key = None


def _get_desc_index(path: Path) -> dict:
    """Return a cached {ItemCode.upper(): description} dict for the given file.

    The source sheet rarely changes while the app is running, so cache by
    workbook path + source sheet name and rebuild only on explicit invalidation.
    """
    global _desc_index_cache, _desc_index_cache_key
    if not path.exists():
        return {}

    cache_key = (str(path.resolve()), str(cfg.SOURCE_SHEET_NAME))
    if cache_key != _desc_index_cache_key:
        _desc_index_cache = _build_description_index(path)
        _desc_index_cache_key = cache_key
    return _desc_index_cache


def get_description_for_itemcode(item_code: str) -> str:
    """Return description for an ItemCode from CREXPD01 index, if available."""
    path = Path(cfg.EXCEL_FILE)
    if not path.exists() or not item_code:
        return ""

    desc_index = _get_desc_index(path)
    if not desc_index:
        return ""

    return desc_index.get(str(item_code).strip().upper(), "") or ""


def load_sheet():
    """Return worksheet rows as list[dict], excluding the header row.

    Description is resolved from a Python-side lookup against CREXPD01 so
    that the Search tab always shows the right value, independent of whether
    Excel has recalculated and cached the formula result. The index is built
    once and cached by file mtime, so repeated searches are instant.
    """
    path = Path(cfg.EXCEL_FILE)
    if not path.exists():
        return []

    # Build (or reuse) the description index BEFORE opening the Heat Number
    # sheet.  Both sheets live in the same ZIP, and opening one sheet's XML
    # stream mid-iteration of another can cause the read-only ZipFile handle
    # to return empty data for the second sheet.
    desc_index = _get_desc_index(path)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = _get_form_sheet_for_read(wb)
    if ws is None:
        wb.close()
        return []

    headers = [c["name"] for c in COLUMNS]
    file_link_col_index = next(
        (idx for idx, col in enumerate(COLUMNS) if col["name"] == "FileLink"), None
    )

    min_data_row = APP_DATA_START_ROW
    max_data_row = ws.max_row

    rows = []
    saw_data = False
    consecutive_empty = 0
    for row in ws.iter_rows(min_row=min_data_row, max_row=max_data_row):
        row_dict = {}
        for i in range(len(headers)):
            cell = row[i] if i < len(row) else None
            if cell is None:
                row_dict[headers[i]] = None
                continue
            hyperlink = getattr(cell, "hyperlink", None)
            if file_link_col_index is not None and i == file_link_col_index and hyperlink:
                row_dict[headers[i]] = getattr(hyperlink, "target", None) or cell.value
            else:
                row_dict[headers[i]] = cell.value

        if any(v not in (None, "") for v in row_dict.values()):
            saw_data = True
            consecutive_empty = 0
            if row_dict.get("ItemCode") and desc_index:
                ic = str(row_dict["ItemCode"]).strip().upper()
                looked_up = desc_index.get(ic, "")
                if looked_up:
                    row_dict["Description"] = looked_up
            rows.append(row_dict)
        else:
            consecutive_empty += 1
            if consecutive_empty > EMPTY_ROW_STOP_THRESHOLD:
                break

    wb.close()
    return rows


def _excel_recalc_and_save(file_path: Path) -> None:
    """Open the workbook in Excel via PowerShell COM, recalculate, and save.

    This caches all formula results as <v> elements in the XML so that
    openpyxl data_only=True can read them immediately on the next load.
    Runs silently in the background — errors are suppressed so a missing
    Excel installation doesn't break the insert flow.
    """
    ps = (
        "$xl = New-Object -ComObject Excel.Application;"
        "$xl.Visible = $false; $xl.DisplayAlerts = $false;"
        "$wb = $xl.Workbooks.Open($env:_EXCEL_PATH);"
        "$xl.CalculateFull(); $wb.Save(); $wb.Close($false); $xl.Quit()"
    )
    try:
        subprocess.run(
            ["powershell", "-NonInteractive", "-NoProfile", "-Command", ps],
            env={**os.environ, "_EXCEL_PATH": str(file_path.absolute())},
            timeout=30,
            capture_output=True,
        )
    except Exception:
        pass  # non-fatal: description will update next time Excel saves


def recalc_workbook() -> None:
    """Force Excel recalculation/save for the configured workbook."""
    file_path = Path(cfg.EXCEL_FILE)
    if not file_path.exists():
        return
    _excel_recalc_and_save(file_path)
    _invalidate_desc_index_cache()


def append_row(data: dict):
    """Append one row to the Heat Number sheet via direct ZIP surgery.

    Phase 1 — read-only streaming pass to find the true last data row.
    Phase 2 — rewrite only the Heat Number worksheet XML inside the ZIP,
               leaving CREXPD01 untouched (never parsed, never serialized).
    """
    file_path = Path(cfg.EXCEL_FILE)
    if not file_path.exists():
        raise FileNotFoundError(f"Workbook not found: {cfg.EXCEL_FILE}")
    try:
        headers = [col["name"] for col in COLUMNS]

        wb = load_workbook(file_path, read_only=True, data_only=True, keep_vba=file_path.suffix.lower() == ".xlsm")
        try:
            ws_form = _get_form_sheet_for_read(wb)
            if ws_form is None:
                last_data_row = APP_DATA_START_ROW - 1
            else:
                last_data_row = APP_DATA_START_ROW - 1
                saw_data = False
                empty_count = 0
                for row_idx, row in enumerate(
                    ws_form.iter_rows(min_row=APP_DATA_START_ROW, max_row=ws_form.max_row, values_only=True),
                    start=APP_DATA_START_ROW,
                ):
                    row_vals = row[:len(headers)] if row else ()
                    if any(v not in (None, "") for v in row_vals):
                        saw_data = True
                        empty_count = 0
                        last_data_row = row_idx
                    else:
                        empty_count += 1
                        if saw_data and empty_count > EMPTY_ROW_STOP_THRESHOLD:
                            break
                        if not saw_data and empty_count > EMPTY_ROW_STOP_THRESHOLD:
                            break
        finally:
            wb.close()

        new_row_idx = max(last_data_row + 1, APP_DATA_START_ROW)

        row_values = [data.get(name, "") for name in headers]
        cached_values = {}

        desc_col_idx = next((i for i, name in enumerate(headers) if name == "Description"), None)
        item_code_col_idx = next((i for i, name in enumerate(headers) if name == "ItemCode"), None)

        if desc_col_idx is not None and item_code_col_idx is not None:
            item_code_col_letter = _col_idx_to_letter(item_code_col_idx)
            item_code_value = str(data.get("ItemCode", "") or "").strip()
            row_values[desc_col_idx] = build_description_formula(
                f"{item_code_col_letter}{new_row_idx}", cfg.SOURCE_SHEET_NAME
            )
            # Seed formula cached value so Search/Insert see Description immediately
            # even before Excel performs a recalc/save pass.
            if item_code_value:
                looked_up = get_description_for_itemcode(item_code_value)
                if looked_up:
                    cached_values[desc_col_idx] = looked_up

        _zip_append_row(file_path, row_values, new_row_idx, cached_values=cached_values)
    except PermissionError as exc:
        raise PermissionError(
            f"Cannot save workbook. Close '{cfg.EXCEL_FILE}' in Excel and try again."
        ) from exc

    # Return the physical worksheet row number that was written.
    return new_row_idx


def sync_form_sheet_columns(old_columns, new_columns):
    """Sync Heat Number sheet schema to match new columns.

    Existing data is preserved by matching values via old column names.
    Columns removed from config are removed from the sheet.
    Newly added columns are created with blank values for existing rows.
    """
    file_path = Path(cfg.EXCEL_FILE)

    wb = _open_workbook()
    _, ws_form = _get_layout_sheets(wb)

    old_names = [c.get("name") for c in old_columns if c.get("name")]
    new_names = [c.get("name") for c in new_columns if c.get("name")]

    if not new_names:
        raise ValueError("Cannot sync workbook: no target columns defined")

    # Prefer real sheet headers when present, otherwise fall back to old config names.
    sheet_headers = []
    if ws_form.max_row >= 1:
        for col_idx in range(1, ws_form.max_column + 1):
            val = ws_form.cell(row=1, column=col_idx).value
            sheet_headers.append(str(val).strip() if val is not None else "")

    effective_old_names = [h for h in sheet_headers if h] or old_names

    # Capture existing rows as dictionaries keyed by old headers.
    existing_row_dicts = []
    if ws_form.max_row >= 2 and effective_old_names:
        for row_idx in range(2, ws_form.max_row + 1):
            row_dict = {}
            has_data = False
            for col_idx, header in enumerate(effective_old_names, start=1):
                if not header:
                    continue
                value = ws_form.cell(row=row_idx, column=col_idx).value
                row_dict[header] = value
                if value not in (None, ""):
                    has_data = True
            if has_data:
                existing_row_dicts.append(row_dict)

    # Rebuild the sheet with the new schema.
    ws_form.delete_rows(1, ws_form.max_row)
    ws_form.append(new_names)

    desc_col_idx_1based = None
    item_code_col_idx_1based = None
    if "Description" in new_names:
        desc_col_idx_1based = new_names.index("Description") + 1
    if "ItemCode" in new_names:
        item_code_col_idx_1based = new_names.index("ItemCode") + 1

    for row_dict in existing_row_dicts:
        ws_form.append([row_dict.get(name, "") for name in new_names])

        # Ensure existing rows keep a correct Description formula after column
        # reordering. This prevents stale references to old ItemCode columns.
        if desc_col_idx_1based is not None and item_code_col_idx_1based is not None:
            row_num = ws_form.max_row
            item_code_cell_ref = f"{_col_idx_to_letter(item_code_col_idx_1based - 1)}{row_num}"
            ws_form.cell(
                row=row_num,
                column=desc_col_idx_1based,
                value=build_description_formula(item_code_cell_ref, cfg.SOURCE_SHEET_NAME),
            )

    try:
        wb.save(file_path)
    finally:
        wb.close()

    # Schema sync can change formulas/references; reset in-memory indexes so
    # Insert/Search rebuild their cached lookups against the new workbook shape.
    _invalidate_desc_index_cache()


def update_file_link(file_number: str, file_link: str) -> bool:
    """Update FileLink for a row in Heat Number sheet identified by File Number.

    Returns True when a matching row was updated, else False.
    """
    if not file_number:
        return False

    file_path = Path(cfg.EXCEL_FILE)
    if not file_path.exists():
        raise FileNotFoundError(f"Workbook not found: {cfg.EXCEL_FILE}")

    wb = _open_workbook()
    try:
        _, ws_form = _get_layout_sheets(wb)

        # Use app schema instead of worksheet row headers. The workbook can
        # contain templates/merged headers, but COLUMNS is the source of truth.
        column_names = [col["name"] for col in COLUMNS]
        try:
            file_number_idx = column_names.index("File Number") + 1
            file_link_idx = column_names.index("FileLink") + 1
        except ValueError as exc:
            raise ValueError("Required columns 'File Number' and/or 'FileLink' are missing in app schema.") from exc

        target = str(file_number).strip().upper()
        updated = False

        for row_idx in range(APP_DATA_START_ROW, ws_form.max_row + 1):
            current_value = ws_form.cell(row=row_idx, column=file_number_idx).value
            if str(current_value or "").strip().upper() != target:
                continue

            link_cell = ws_form.cell(row=row_idx, column=file_link_idx)
            link_cell.value = file_link
            if file_link:
                link_cell.hyperlink = file_link
                link_cell.style = "Hyperlink"
            else:
                link_cell.hyperlink = None

            updated = True
            break

        if updated:
            wb.save(file_path)
        return updated
    finally:
        wb.close()


def search_rows(search_value, search_column="ItemCode"):
    rows = load_sheet()
    if not search_value:
        return rows

    search_value = str(search_value).strip().lower()
    column = search_column if search_column in SEARCH_BY else "ItemCode"

    results = []
    for row in rows:
        candidate = str(row.get(column, "")).strip().lower()
        if search_value in candidate:
            results.append(row)
    return results
